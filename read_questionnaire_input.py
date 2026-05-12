#!/usr/bin/env python3
"""Flatten HPWinner intake questionnaire .xlsx to JSON for agents.

Supports:

- **v2** (`HPWinner_Intake_Questionnaire_v2.xlsx`): sheets ``中文_填写`` / ``EN_Form``,
  header row 1, answers in column **E** (after 必填 / Required in D).
- **Legacy**: sheets ``填写`` / ``English``, answers in column **D**.

The answer column is detected from row 1 (looks for "回答" or "answer").

Default input path (when run from monorepo root): questionnaire/questionnaire_01_input.xlsx

Usage::

    python3 read_questionnaire_input.py --input HPWinner_Intake_Questionnaire_v2.xlsx
    python3 read_questionnaire_input.py --input path/to/file.xlsx --sheet 中文_填写
    python3 read_questionnaire_input.py --answers-only --compact
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_QID_RE = re.compile(r"^[A-Z]+\d+[a-z]?$")


def _json_safe(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return str(value)


def _answer_column(ws) -> int:
    """1-based column index for the answer cell (legacy: D=4, v2: E=5)."""
    raw4 = ws.cell(1, 4).value
    raw5 = ws.cell(1, 5).value
    s4 = str(raw4).strip() if raw4 is not None else ""
    s5 = str(raw5).strip() if raw5 is not None else ""
    low4, low5 = s4.lower(), s5.lower()

    if "回答" in s5 or low5 == "answer" or low5.startswith("answer"):
        return 5
    if "回答" in s4 or low4 == "answer" or low4.startswith("answer"):
        return 4
    return 4


def _looks_like_question_id(qid: str) -> bool:
    return bool(_QID_RE.match(qid.strip()))


def _read_sheet(ws) -> list[dict[str, Any]]:
    ans_col = _answer_column(ws)
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        qid_raw = ws.cell(r, 1).value
        if qid_raw is None or str(qid_raw).strip() == "":
            continue
        qid = str(qid_raw).strip()
        if not _looks_like_question_id(qid):
            continue
        rows.append(
            {
                "id": qid,
                "question": _json_safe(ws.cell(r, 2).value),
                "unit": _json_safe(ws.cell(r, 3).value),
                "answer": _json_safe(ws.cell(r, ans_col).value),
            }
        )
    return rows


def default_both_sheets(sheetnames: list[str]) -> list[str]:
    """Prefer v2 pair, then legacy pair; omit missing half."""
    out: list[str] = []
    if "中文_填写" in sheetnames:
        out.append("中文_填写")
    if "EN_Form" in sheetnames:
        out.append("EN_Form")
    if out:
        return out
    if "填写" in sheetnames:
        out.append("填写")
    if "English" in sheetnames:
        out.append("English")
    if out:
        return out
    raise ValueError(
        "No known data sheets. Expected "
        "「中文_填写」/「EN_Form」 (v2) or 「填写」/「English」 (legacy). "
        f"Found: {sheetnames}"
    )


def flatten_workbook(path: Path, sheets: list[str]) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    out: dict[str, Any] = {"source": str(path.resolve()), "sheets": {}}
    try:
        for name in sheets:
            if name not in wb.sheetnames:
                raise ValueError(f"Sheet {name!r} not in workbook; have: {wb.sheetnames}")
            out["sheets"][name] = {"rows": _read_sheet(wb[name])}
    finally:
        wb.close()
    return out


def answers_dict(flat: dict[str, Any], sheet_name: str) -> dict[str, Any]:
    d: dict[str, Any] = {}
    for row in flat["sheets"][sheet_name]["rows"]:
        qid = row["id"]
        if qid in d:
            sys.stderr.write(f"warning: duplicate id {qid!r} in sheet {sheet_name!r}; using last\n")
        d[qid] = row["answer"]
    return d


def primary_answers_sheet(flat: dict[str, Any]) -> str:
    keys = flat["sheets"]
    if "中文_填写" in keys:
        return "中文_填写"
    if "填写" in keys:
        return "填写"
    return next(iter(keys))


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--input",
        "-i",
        type=Path,
        default=None,
        help="Path to questionnaire .xlsx (default: questionnaire/questionnaire_01_input.xlsx under cwd)",
    )
    p.add_argument(
        "--sheet",
        "-s",
        default="both",
        metavar="NAME",
        help='Export sheet(s): "both" (default: v2 or legacy pair), or one of: '
        "中文_填写, EN_Form, 填写, English",
    )
    p.add_argument(
        "--answers-only",
        action="store_true",
        help="Emit only { id: answer } from the primary sheet (中文_填写 or 填写 when present)",
    )
    p.add_argument(
        "--compact",
        action="store_true",
        help="Minified JSON (no indent)",
    )
    args = p.parse_args(argv)

    path = args.input
    if path is None:
        path = Path("questionnaire/questionnaire_01_input.xlsx")
    path = path.expanduser().resolve()
    if not path.is_file():
        print(f"error: file not found: {path}", file=sys.stderr)
        return 1

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()

    if args.sheet == "both":
        try:
            sheets = default_both_sheets(names)
        except ValueError as e:
            print(f"error: {e}", file=sys.stderr)
            return 1
    else:
        if args.sheet not in names:
            print(f"error: sheet {args.sheet!r} not in workbook; have: {names}", file=sys.stderr)
            return 1
        sheets = [args.sheet]

    try:
        flat = flatten_workbook(path, sheets)
    except ValueError as e:
        print(f"error: {e}", file=sys.stderr)
        return 1

    if args.answers_only:
        payload = answers_dict(flat, primary_answers_sheet(flat))
    else:
        payload = flat

    indent = None if args.compact else 2
    json.dump(payload, sys.stdout, ensure_ascii=False, indent=indent)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
